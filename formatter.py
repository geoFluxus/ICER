import warnings
import pandas as pd
import openpyxl

warnings.simplefilter(action='ignore', category=FutureWarning)
pd.options.display.width = 0
pd.set_option('display.max_rows', None)

"""This script converts the dummy CBS file into the standard CBS data template"""

corop_to_province = {
'Oost-Groningen': 'Groningen',
'Delfzijl e.o.': 'Groningen',
'Overig Groningen': 'Groningen',
'Noord-Friesland': 'Friesland',
'Zuidwest-Friesland': 'Friesland',
'Zuidoost-Friesland': 'Friesland',
'Noord-Drenthe': 'Drenthe',
'Zuidoost-Drenthe': 'Drenthe',
'Zuidwest-Drenthe': 'Drenthe',
'Noord-Overijssel': 'Overijssel',
'Zuidwest-Overijssel': 'Overijssel',
'Twente': 'Overijssel',
'Veluwe': 'Gelderland',
'Achterhoek': 'Gelderland',
'Aggl. Arnhem/Nijmegen': 'Gelderland',
'Zuidwest-Gelderland': 'Gelderland',
'Utrecht-West': 'Utrecht',
'Stadsgewest Amersfoort': 'Utrecht',
'Stadsgewest Utrecht': 'Utrecht',
'Zuidoost-Utrecht': 'Utrecht',
'Kop van Noord-Holland': 'Noord-Holland',
'Alkmaar e.o.': 'Noord-Holland',
'IJmond': 'Noord-Holland',
'Agglomeratie Haarlem': 'Noord-Holland',
'Overig Agglomeratie Amsterdam': 'Noord-Holland',
'Zaanstreek': 'Noord-Holland',
'Haarlemmermeer e.o.': 'Noord-Holland',
'Groot-Amsterdam': 'Noord-Holland',
'Amsterdam': 'Noord-Holland',
'Edam-Volendam e.o.': 'Noord-Holland',
'Het Gooi en Vechtstreek': 'Noord-Holland',

'Agglomeratie Leiden en Bollenstreek': 'Zuid-Holland',
r"Agglomeratie’s-Gravenhage (Excl. Zoetermeer)": "Zuid-Holland",
'Zoetermeer': 'Zuid-Holland',
'Delft en Westland': 'Zuid-Holland',
'Oost-Zuid-Holland': 'Zuid-Holland',
'Rijnmond': 'Zuid-Holland',
'Overig Groot-Rijnmond': 'Zuid-Holland',
'Drechtsteden': 'Zuid-Holland',
'Overig Zuidoost-Zuid-Holland': 'Zuid-Holland',
'Zeeuwsch-Vlaanderen': "Zeeland",
'Overig Zeeland': "Zeeland",
'West-Noord-Brabant': "Noord-Brabant",
'Midden-Noord-Brabant': "Noord-Brabant",
r"Stadsgewest ’s-Hertogenbosch": 'Noord-Brabant',
'Overig Noordoost-Noord-Brabant': "Noord-Brabant",
'Zuidoost-Noord-Brabant': "Noord-Brabant",
'Noord-Limburg': "Limburg",
'Midden-Limburg': "Limburg",
'Zuid-Limburg': "Limburg",
'Almere': "Flevoland",
'Flevoland-Midden': "Flevoland",
'Noordoostpolder en Urk': "Flevoland"
}

def correct_data_gaps(data):
    # Filter the missing data for year 2022 and Suikerbieten
    print(data.columns)
    #print(data)
    missing_data = [data[(data["Jaar"] == 2022) & (data["Goederengroep_naam"] == "Suikerbieten")],
                    data[(data["Jaar"] == 2015) & (data["Goederengroep_naam"] == "Rauwe melk van runderen, schapen en geiten")],
                    data[(data["Jaar"] == 2022) &
                         (data["Goederengroep_naam"] == "Ruwe aardolie") &
                         (data['Provincienaam'].isin(['Zuid-Holland', 'Noord-Brabant']))],

                    ]


    # Loop through each missing entry and fill it with the previous year's value
    for dat in missing_data:
        for index, row in dat.iterrows():
            # Find the previous year's value for the same region and good group
            year = 2021 if row['Jaar'] == 2022 else 2016
            previous_year_value = data[
                (data["Jaar"] == year) &
                (data["Provincienaam"] == row["Provincienaam"]) &
                (data["Goederengroep_naam"] == row["Goederengroep_naam"]) &
                 (data['Stroom'] == row['Stroom'])
                ][['Brutogew', 'Sf_brutogew', 'Waarde', 'Sf_waarde']].values

            # If a previous year value exists, update the missing value
            vals = ['Brutogew', 'Sf_brutogew', 'Waarde', 'Sf_waarde']
            if len(previous_year_value) > 0:
                for i in range(len(vals)):
                    data.at[index, vals[i]] = previous_year_value[0][i]
    print(data)

    return data

def run(data, filename, corop=False, fill_data_gaps=False):

    # filter out totals from the gebruiksgroep_naam, group by
    data = data[data['Gebruiksgroep_naam'] != 'Totaal']

    # drop unnecessary columns
    if not corop:
        data = data.drop(columns=['Stroom_nr', 'Provincie_nr', 'Gebruiksgroep_nr', 'Gebruiksgroep_naam'])

        #If there are columns left empty; typecast all values to floats
        for i in ['Brutogew', 'Sf_brutogew', 'Waarde', 'Sf_waarde']:
            data[i] = data[i].str.replace(',', '.')
            data[i] = data[i].str.replace(' ', '0')
            data[i] = data[i].astype(float)
    else:
        for i in ['Brutogew', 'Sf_brutogew', 'Waarde', 'Sf_waarde']:
            data[i] = data[i].str.replace(',', '.')
            data[i] = data[i].str.replace(' ', '0')
            data[i] = data[i].astype(float)
        data['Provincienaam'] = data['Regionaam'].apply(lambda x: corop_to_province[x])
        data = data.drop(columns=['Stroom_nr', 'Regionr', 'Gebruiksgroep_nr', 'Gebruiksgroep_naam'])


    # group by
    data = data.groupby(['Jaar', 'Stroom', 'Provincienaam', 'Goederengroep_nr', 'Goederengroep_naam']).sum().reset_index()

    if fill_data_gaps:
        data = correct_data_gaps(data)
    cols = {'gewicht': ['Brutogew', 'Sf_brutogew'],
            'waarde': ['Waarde', 'Sf_waarde']}
    text = ' Aangepast' if fill_data_gaps else ''

    # Create an Excel writer object
    with pd.ExcelWriter(f'data/{filename}{text}.xlsx', engine='openpyxl') as writer:
        no = 0
        letter = 'b'
        # split data in data sheets according to the format
        for year in data['Jaar'].unique():
            data_year = data[data['Jaar'] == year]
            no += 1

            for tab in ['gewicht', 'waarde']:
                cols_tab = cols[tab]
                data_tab = data_year[['Jaar', 'Stroom', 'Provincienaam', 'Goederengroep_nr', 'Goederengroep_naam'] + cols_tab]

                data_tab = data_tab.rename(columns={'Goederengroep_naam': 'Goederengroep',
                                                    'Provincienaam': 'Provincie',
                                                    cols_tab[0]: tab,
                                                    cols_tab[1]: 'standaard-fout'})


                # Pivoting the DataFrame
                data_pivoted = data_tab.pivot_table(
                    index=['Jaar', 'Provincie', 'Goederengroep', 'Goederengroep_nr'],
                    columns='Stroom',
                    values=[tab, 'standaard-fout'],
                    aggfunc='first'  # Choose 'first' to pick first value in case of duplicates, or modify as needed
                )

                # Flatten the column MultiIndex created by pivot_table
                data_pivoted.columns = [f'{col[1]} {col[0]}' for col in data_pivoted.columns]

                # Resetting index to turn it back into a DataFrame
                data_tab = data_pivoted.reset_index()

                data_tab.drop(columns='Jaar', inplace=True)
                # Format sheet name
                if tab == 'gewicht':
                    letter = 'a'
                    unit = 'kg'
                elif tab == 'waarde':
                    letter = 'b'
                    unit = 'euro'
                sheet_name = f'Tabel {no}{letter}'

                # Write the dataframe to a sheet starting from the fourth row (index 3)
                data_tab.to_excel(writer, sheet_name=sheet_name, startrow=3, header=False, index=False)

                # Get the workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Write the custom text in the first and third rows
                worksheet.cell(row=1, column=1, value=f"Tabel {no}{letter}. Brutogewicht van in-, uit-, wederuit- en doorvoer, aanbod eigen regio en distributie naar provincie en goederengroep, {year}")
                worksheet.cell(row=3, column=3,
                               value=f"x mln {unit}")

                # Write the column names in the second row
                for idx, col in enumerate(data_tab.columns, 1):
                    worksheet.cell(row=2, column=idx, value=col)



if __name__ == '__main__':
    filepath = 'data/'
    filename = 'CBS/041224 Tabel Regionale stromen 2015-2022 provincie CE67 GC6'

    all_data = pd.read_csv(filepath + filename +'.csv', delimiter=';', decimal=',', encoding='cp1252')
    run(all_data, filename, fill_data_gaps=True)